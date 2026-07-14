import base64
import io
import json

from odoo import api, fields, models
from odoo.exceptions import UserError

from .transforms import apply_transform


class PayrollImportBatch(models.Model):
    """One uploaded file, its whole life: draft -> parsed -> validated (dry-run) -> committed | error.
    Never deleted — it is the audit trail. Commit is an idempotent upsert in a savepoint; any failure
    rolls the batch back and names the failing row."""

    _name = "payroll.import.batch"
    _description = "Payroll Import Batch"
    _inherit = ["mail.thread"]
    _order = "create_date desc"

    name = fields.Char(required=True, default="New import", tracking=True)
    profile_id = fields.Many2one("payroll.import.profile", required=True)
    client_id = fields.Many2one(related="profile_id.client_id", store=True)
    period = fields.Char(help="Overrides the profile period for this batch.")
    data_file = fields.Binary(required=True, attachment=True)
    filename = fields.Char()
    state = fields.Selection(
        [("draft", "Draft"), ("parsed", "Parsed"), ("validated", "Validated"),
         ("committed", "Committed"), ("error", "Error")],
        default="draft", required=True, tracking=True)
    line_ids = fields.One2many("payroll.import.line", "batch_id")
    error_count = fields.Integer(compute="_compute_counts")
    row_count = fields.Integer(compute="_compute_counts")

    @api.depends("line_ids", "line_ids.error")
    def _compute_counts(self):
        for b in self:
            b.row_count = len(b.line_ids)
            b.error_count = len(b.line_ids.filtered("error"))

    # ---------------------------------------------------------------- parse
    def action_parse(self):
        self.ensure_one()
        self.line_ids.unlink()
        rows = self._read_rows()
        profile = self.profile_id
        mapping = profile.mapping_dict()
        header_row = profile.header_row or 1
        start = (profile.data_start_row or (header_row + 1)) - 1
        headers = [str(h).strip() if h is not None else "" for h in rows[header_row - 1]]
        vals_list = []
        for i, row in enumerate(rows[start:], start=start):
            joined = "".join(str(c) for c in row if c is not None)
            if not joined or "section break" in joined.lower() or "total" in joined.lower():
                continue  # skip blank rows, section breaks and trailing TOTAL rows
            resolved, emp_code = {}, False
            for col, header in enumerate(headers):
                ml = mapping.get(header)
                if not ml:
                    continue
                raw = row[col] if col < len(row) else None
                value = apply_transform(ml.transform, raw, ml.transform_arg)
                if ml.target_field == "emp_code":
                    emp_code = value
                else:
                    resolved[ml.target_field] = value
            if not emp_code and not resolved:
                continue
            vals_list.append({
                "batch_id": self.id, "row_index": i + 1, "emp_code": emp_code or False,
                "raw_json": json.dumps([str(c) if c is not None else "" for c in row]),
                "values_json": json.dumps(resolved),
            })
        self.env["payroll.import.line"].create(vals_list)
        self.state = "parsed"
        self.message_post(body="Parsed %d data rows." % len(vals_list))

    def _read_rows(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError("openpyxl is required to read .xlsx files.")
        if not self.data_file:
            raise UserError("Upload a file first.")
        data = base64.b64decode(self.data_file)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb[self.profile_id.sheet_name] if self.profile_id.sheet_name else wb[wb.sheetnames[0]]
        off = self.profile_id.col_offset or 0
        return [tuple(r)[off:] for r in ws.iter_rows(values_only=True)]

    # ---------------------------------------------------------------- validate (dry-run)
    def action_validate(self):
        self.ensure_one()
        if self.state == "draft":
            self.action_parse()
        Employee = self.env["hr.employee"]
        errs = 0
        for line in self.line_ids:
            problems = []
            if not line.emp_code:
                problems.append("missing employee code")
            elif not Employee.search_count([
                    ("barcode", "=", line.emp_code), ("company_id", "=", self.client_id.id)]):
                problems.append("no employee %s in %s" % (line.emp_code, self.client_id.name))
            line.error = "; ".join(problems) or False
            errs += bool(problems)
        self.state = "validated"
        self.message_post(body="Validated: %d rows, %d exception(s). Nothing written." % (
            len(self.line_ids), errs))

    # ---------------------------------------------------------------- commit (idempotent)
    def action_commit(self):
        self.ensure_one()
        if self.state != "validated":
            self.action_validate()
        if self.error_count:
            raise UserError(
                "%d row(s) have exceptions. Fix the file and re-import; nothing was written."
                % self.error_count)
        period = self.period or self.profile_id.period
        if not period:
            raise UserError("No period set on the batch or profile.")
        MonthlyInput = self.env["payroll.monthly.input"]
        try:
            with self.env.cr.savepoint():
                for line in self.line_ids:
                    values = json.loads(line.values_json or "{}")
                    for code, value in values.items():
                        key = [("client_id", "=", self.client_id.id), ("period", "=", period),
                               ("emp_code", "=", line.emp_code), ("input_code", "=", code)]
                        existing = MonthlyInput.search(key, limit=1)
                        if existing:
                            existing.value = value          # upsert: no duplicate on re-import
                        else:
                            MonthlyInput.create({
                                "client_id": self.client_id.id, "period": period,
                                "emp_code": line.emp_code, "input_code": code, "value": value})
        except Exception as exc:
            self.state = "error"
            raise UserError("Commit failed and was rolled back: %s" % exc)
        self.state = "committed"
        self.message_post(body="Committed %d rows to %s (idempotent upsert)." % (len(self.line_ids), period))


class PayrollImportLine(models.Model):
    _name = "payroll.import.line"
    _description = "Payroll Import Staging Line"
    _order = "row_index"

    batch_id = fields.Many2one("payroll.import.batch", required=True, ondelete="cascade")
    row_index = fields.Integer()
    emp_code = fields.Char(index=True)
    raw_json = fields.Text(string="Raw row")
    values_json = fields.Text(string="Resolved values")
    error = fields.Char()
