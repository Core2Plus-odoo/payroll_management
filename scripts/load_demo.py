#!/usr/bin/env python3
"""
Idempotent demo loader — populates employees, contracts (hr.version) and monthly inputs for the four
EY clients from the canonical workbook. Run inside the Odoo.sh / dev shell:

    odoo-bin shell < scripts/load_demo.py            # or:  make demo

Static config (companies, structures, salary rules, import profiles, statutory rates) is installed by
the c2p_payroll_demo module data. This loader adds the bulk people data on top, and can be re-run
safely — employees are matched by barcode (= employee code), inputs are upserted.
"""
import logging
import os

_logger = logging.getLogger("load_demo")

XLSX = os.path.join(os.path.dirname(__file__), "..", "source_data", "EY_Payroll_Master_Data.xlsx")
PERIOD = "2026-06"
PROVINCE = {"ITM": "Sindh", "FSL": "Punjab", "MBS": "ICT", "GRL": False}
COMPANY_NAME = {
    "ITM": "Indus Textile Mills (Pvt) Ltd", "FSL": "Falcon Software (Pvt) Ltd",
    "MBS": "Meridian BPO Services (Pvt) Ltd", "GRL": "Gulf Retail LLC",
}
STRUCT_XMLID = {
    "ITM": "c2p_payroll_demo.struct_ss_itm_factory", "FSL": "c2p_payroll_demo.struct_ss_fsl_whitecollar",
    "MBS": "c2p_payroll_demo.struct_ss_mbs_callcentre", "GRL": "c2p_payroll_demo.struct_ss_grl_retail",
}
STYPE_XMLID = {
    "ITM": "c2p_l10n_pk_hr_payroll.structure_type_pk_monthly",
    "FSL": "c2p_l10n_pk_hr_payroll.structure_type_pk_monthly",
    "MBS": "c2p_l10n_pk_hr_payroll.structure_type_pk_monthly",
    "GRL": "c2p_payroll_demo.structure_type_uae_monthly",
}


def load(env):
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    companies = {code: env["res.company"].search([("name", "=", name)], limit=1)
                 for code, name in COMPANY_NAME.items()}
    Employee = env["hr.employee"]
    MonthlyInput = env["payroll.monthly.input"]

    # ---- employees + contract/version (wage = monthly gross) ----
    emp_rows = list(wb["07_employees"].iter_rows(values_only=True))
    con = {r[0]: r for r in wb["08_contracts"].iter_rows(values_only=True)}
    n_emp = 0
    for r in emp_rows:
        emp_code, client = r[0], r[1]
        if client not in companies or not companies[client]:
            continue
        contract = con.get(emp_code)
        if not contract:
            continue
        wage = contract[7] or 0.0
        vals = {
            "name": r[4], "company_id": companies[client].id, "barcode": emp_code,
            "pk_pf_member": (r[19] == "yes"), "pk_cnic": r[10], "payroll_province": PROVINCE[client],
        }
        employee = Employee.with_company(companies[client]).search([("barcode", "=", emp_code)], limit=1)
        if employee:
            employee.write(vals)
        else:
            employee = Employee.with_company(companies[client]).create(vals)
        version = employee.version_id
        version.write({
            "wage": wage,
            "contract_date_start": str(contract[5]) if contract[5] else "2019-01-01",
            "structure_type_id": env.ref(STYPE_XMLID[client]).id,
        })
        n_emp += 1

    # ---- monthly variable inputs (upsert -> idempotent) ----
    codes = ["OT_H", "ABSENT_D", "BONUS", "COMM", "SHIFT_ALW", "OTH_DED"]
    n_inp = 0
    for r in wb["09_monthly_inputs"].iter_rows(values_only=True):
        emp_code, client = r[0], r[1]
        if client not in companies or not companies[client]:
            continue
        for i, code in enumerate(codes):
            value = r[3 + i] or 0.0
            if not value:
                continue
            key = [("client_id", "=", companies[client].id), ("period", "=", PERIOD),
                   ("emp_code", "=", emp_code), ("input_code", "=", code)]
            existing = MonthlyInput.search(key, limit=1)
            if existing:
                existing.value = value
            else:
                MonthlyInput.create({
                    "client_id": companies[client].id, "period": PERIOD,
                    "emp_code": emp_code, "input_code": code, "value": value})
            n_inp += 1

    env.cr.commit()
    _logger.info("load_demo: %d employees, %d monthly inputs loaded for %s", n_emp, n_inp, PERIOD)
    print("load_demo: %d employees, %d monthly inputs loaded for %s" % (n_emp, n_inp, PERIOD))


# When piped into `odoo-bin shell`, `env` is in scope.
if "env" in dir():
    load(env)  # noqa: F821
